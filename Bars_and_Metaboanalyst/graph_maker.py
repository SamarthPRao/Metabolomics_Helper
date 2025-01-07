import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
from matplotlib.ticker import MaxNLocator
from openpyxl import Workbook

def create_bar_graph(data, title, single_color="#78A2CC"):
    categories = data.index
    values = data.values

    fontsize = 12 + len(categories) * 0.1

    sns.set(style="white")
    fig, ax = plt.subplots(figsize=(len(categories) * 0.15 + 6, len(categories) * 0.3 + 3))  # Adjust size based on data
    sns.barplot(y=categories, x=values, color=single_color, ax=ax)

    for index, value in enumerate(values):
        ax.text(value + 0.2 + max(values) * 0.01, index, str(value), color='black', va='center', fontsize=fontsize)

    ax.grid(False)
    sns.despine()
    ax.spines['bottom'].set_visible(False)

    plt.xticks(fontsize=fontsize)
    plt.yticks(fontsize=fontsize)
    ax.set_ylabel('')

    # Ensure x-ticks are integers
    ax.xaxis.set_major_locator(MaxNLocator(integer=True))

    plt.tight_layout()

    return fig  # Return the figure object


def generate_graph(file):
    df = pd.read_csv(file)
    df = df[['Main Category', 'General Classification', 'Sub-class']]

    # Data preparation
    data1 = df['Main Category'].value_counts()
    data2 = df['General Classification'].value_counts()
    data3 = df[df['General Classification'] == 'Flavonoid']['Sub-class'].value_counts()

    # Create bar graphs
    figures = []
    figures.append(create_bar_graph(data1, 'Main Categories'))
    figures.append(create_bar_graph(data2, 'Sub Categories'))
    figures.append(create_bar_graph(data3, 'Flavonoid Sub-Categories'))

    # Create Excel workbook
    wb = Workbook()

    # Sheet 1: Main Categories
    ws = wb.active
    ws.title = 'Main Categories'
    ws.append(['Main Category', 'Count'])
    for category, count in data1.items():
        ws.append([category, count])

    # Sheet 2: Sub Categories
    wb.create_sheet('Sub Categories')
    ws = wb.worksheets[1]
    ws.append(['Sub Category', 'Count'])
    for category, count in data2.items():
        ws.append([category, count])

    # Sheet 3: Flavonoid Sub Categories
    wb.create_sheet('Flavonoid Sub Categories')
    ws = wb.worksheets[2]
    ws.append(['Sub Category', 'Count'])
    for category, count in data3.items():
        ws.append([category, count])

    return figures, wb  # Return figures and workbook