import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.formatting.rule import FormulaRule

# Function for removing nan's
def remove_nans(inp):
    out = []
    for i in inp:
        if type(i) == str:
            out.append(i)
    return out

# Function to find compound locations
def get_compound_locations(df, lower_bar):
        compound_locations = []
        for i in range(df.shape[0]):
            if remove_nans(df.iloc[i]) == lower_bar:
                compound_locations.append(i - 1)
        return compound_locations


def process_files(cs, mz, ml, database):

    df_cs = pd.read_excel(cs)
    df_mz = pd.read_excel(mz)
    df_ml = pd.read_excel(ml)

    df_database = pd.read_excel(database)

    # Define the lower bar with headers for compound data
    lower_bar_cs = remove_nans(df_cs.iloc[1])
    lower_bar_mz = remove_nans(df_mz.iloc[1])
    lower_bar_ml = remove_nans(df_ml.iloc[1])
    
    # find where compounds are located in all three sheets
    compound_locations_cs = get_compound_locations(df_cs, lower_bar_cs)
    compound_locations_mz = get_compound_locations(df_mz, lower_bar_mz)
    compound_locations_ml = get_compound_locations(df_ml, lower_bar_ml)


    # Finds relevant indices for compound search later
    cs_id_location, cs_ref_location, cs_structure_location = [df_cs.iloc[1].tolist().index(k) for k in ['CSID', '# References', 'Structure']]
    mz_id_location, mz_match_location, mz_structure_location = [df_mz.iloc[1].tolist().index(k) for k in ['mzCloud ID', 'Best Match', 'Structure']]
    ml_id_location, ml_structure_location = [df_ml.iloc[1].tolist().index(k) for k in ['FL_index', 'Structure']]

    #========================================================================#
    #=====                          ChemSpider                          =====#
    #========================================================================#


    # Find column where 'Match Type' is located, if it exists
    if 'Match Type' in lower_bar_cs:
        cs_match_type_location = df_cs.iloc[1].tolist().index('Match Type')
    else:
        cs_match_type_location = None

    # Prepare output lists and set for fast lookup
    csid_list = []
    cs_structure_list = []
    compound_locations_cs_set = set(compound_locations_cs)

    # Determine the column to use for match type - 'Match Type' may not exist
    match_type_col_cs = (
        cs_match_type_location
        if cs_match_type_location is not None
        else df_cs.columns.get_loc('Annot. Source: ChemSpider Search')
    )

    def get_best_candidate_cs(start_idx, match_col_idx, match_value='Full match'):
        # Return best CSID and structure for a compound starting at start_idx
        candidates = {}
        idx = start_idx
        while idx < df_cs.shape[0] and idx not in compound_locations_cs_set:
            if df_cs.iloc[start_idx - 2, match_col_idx] == match_value:
                ref_count = df_cs.iloc[idx, cs_ref_location]
                candidates[ref_count] = [
                    df_cs.iloc[idx, cs_id_location],
                    df_cs.iloc[idx, cs_structure_location]
                ]
            idx += 1
        return candidates.get(max(candidates, default=0), ['', ''])

    # Process each compound starting point
    for val in np.array(compound_locations_cs) + 2:
        if val in compound_locations_cs_set:
            csid_list.append('')
            cs_structure_list.append('')
        else:
            csid, structure = get_best_candidate_cs(val, match_type_col_cs)
            csid_list.append(csid)
            cs_structure_list.append(structure)

    # Get relevant compound metadata columns
    cols = ['Name', 'Formula', 'Calc. MW', 'RT [min]']
    compound_data = {
        col: df_cs.loc[compound_locations_cs, col].tolist()
        for col in cols
    }

    # Create dictionary while avoiding duplicates
    cs_dict = {}
    for row in zip(
        compound_data['Name'],
        compound_data['Formula'],
        compound_data['Calc. MW'],
        compound_data['RT [min]'],
        csid_list,
        cs_structure_list
    ):
        name = row[0]
        if name not in cs_dict:
            # remember CSID is in number format, so we have to make it a string for link generation later
            cs_dict[name] = list(row[1:-2]) + [str(row[-2]), row[-1]]


    #========================================================================#
    #=====                            mzCloud                           =====#
    #========================================================================#

    mz_id_list = []
    mz_structure_list = []
    compound_locations_mz_set = set(compound_locations_mz)

    # There is no worrying about whether match type exists here
    # So far it is always there in the data we've seen
    for i, val in enumerate(np.array(compound_locations_mz) + 2):
        if val in compound_locations_mz_set:
            mz_id_list.append('')
            mz_structure_list.append('')
        else:
            idx = val
            candidates = {}
            while idx < df_mz.shape[0] and idx not in compound_locations_mz_set:
                if 'Reference' in df_mz.iloc[idx, mz_id_location]:
                    candidates[df_mz.iloc[idx, mz_match_location]] = [df_mz.iloc[idx, mz_id_location][10:],
                                                                      df_mz.iloc[idx, mz_structure_location]]
                idx += 1
            if len(candidates) == 0:
                mz_id_list.append('')
                mz_structure_list.append('')
            else:
                mz_id_list.append(candidates[max(candidates.keys())][0])
                mz_structure_list.append(candidates[max(candidates.keys())][1])
    # Getting compound names for this file
    compound_names_mz = df_mz.loc[:, 'Name'][compound_locations_mz].to_list()

    # create mzCloud dictionary
    mz_name_set = set()
    mz_dict = {}
    for i in range(len(compound_names_mz)):
        if compound_names_mz[i] not in mz_name_set:
            mz_name_set.add(compound_names_mz[i])
            mz_dict[compound_names_mz[i]] = [mz_id_list[i], mz_structure_list[i]]


    #========================================================================#
    #=====                           MassList                           =====#
    #========================================================================#

    # Getting relevant column locations
    ml_match_location = df_ml.iloc[1].tolist().index('Compound Match') if 'Compound Match' in lower_bar_ml else None
    npaid_location = df_ml.iloc[1].tolist().index('npaid') if 'npaid' in lower_bar_ml else None
    match_type_col_ml = df_ml.columns.get_loc('Annot. Source: MassList Search')

    mlid_list = []
    ml_structure_list = []
    compound_locations_ml_set = set(compound_locations_ml)

    # Helper function for ID extraction
    def extract_ml_id(row_idx):
        val = df_ml.iloc[row_idx, ml_id_location]
        # If FL index found
        if isinstance(val, str):
            return val[28:]  # FL##########
        # If npaid index found
        elif npaid_location is not None:
            return df_ml.iloc[row_idx, npaid_location]  # NPA#####
        # If neither found
        return ''

    # Main processing loop
    for val in np.array(compound_locations_ml) + 2:
        if val in compound_locations_ml_set:
            mlid_list.append('')
            ml_structure_list.append('')
            continue # skip to next val

        idx = val
        appended = False
        while idx < df_ml.shape[0] and idx not in compound_locations_ml_set:
            full_match = False

            # Checking for full match depending on the type of data given
            if ml_match_location is not None:
                full_match = df_ml.iloc[idx, ml_match_location] == 'Full match'
            else:
                full_match = df_ml.iloc[val - 2, match_type_col_ml] == 'Full match'

            if full_match:
                mlid_list.append(extract_ml_id(idx))
                ml_structure_list.append(df_ml.iloc[idx, ml_structure_location])
                appended = True
                break
            idx += 1

        if not appended:
            mlid_list.append('')
            ml_structure_list.append('')

    # Build dictionary
    compound_names_ml = df_ml.loc[compound_locations_ml, 'Name'].tolist()
    ml_dict = {}
    for name, mlid, struct in zip(compound_names_ml, mlid_list, ml_structure_list):
        if name not in ml_dict:
            ml_dict[name] = [mlid, struct]

    #========================================================================#
    #=====                       Combining the Data                     =====#
    #========================================================================#

    # Step 1: Start with base info from cs_dict
    output_df = pd.DataFrame.from_dict(cs_dict, orient='index', columns=[
        'Formula', 'Calc. MW', 'RT [min]', 'ChemSpider ID', 'Structure'
    ])
    output_df.index.name = 'Name'
    output_df.reset_index(inplace=True)

    # Step 2: Add mzCloud and MassList IDs using .get() to avoid key errors
    output_df['mzCloud ID'] = output_df['Name'].apply(lambda name: mz_dict.get(name, ['', ''])[0])
    output_df['Mass List ID'] = output_df['Name'].apply(lambda name: ml_dict.get(name, ['', ''])[0])

    # Step 3: Annotation Level
    output_df['Annotation Level'] = output_df[['ChemSpider ID', 'mzCloud ID', 'Mass List ID']].apply(
        lambda row: sum(bool(str(val).strip()) for val in row), axis=1
    )

    # Step 4: Consolidated Structure priority: CS > mzCloud > MassList
    def choose_structure(name):
        if cs_dict.get(name, ['', '', '', '', ''])[4]:
            return cs_dict[name][4]
        if name in mz_dict and mz_dict[name][1]:
            return mz_dict[name][1]
        if name in ml_dict and ml_dict[name][1]:
            return ml_dict[name][1]
        return ''
    output_df['Structure'] = output_df['Name'].apply(choose_structure)

    # Step 5: Add metabolite database info
    # Build metabolite_dict from database
    metabolite_dict = {
        row[1]: [row[0], row[2], row[3], row[4]]
        for row in df_database.itertuples(index=False)
    }

    def get_meta(name):
        return metabolite_dict.get(name, ['', '', '', ''])

    meta_info = output_df['Name'].apply(get_meta)
    output_df[['Main Category', 'General Classification', 'Sub-class', 'Comments']] = pd.DataFrame(meta_info.tolist(), index=output_df.index)

    # Step 6: Add ESI Mode
    output_df['ESI Mode'] = 'ESI+'

    # Step 7: Final column order
    desired_cols = [
        'Main Category', 'Name', 'Formula', 'ESI Mode', 'Calc. MW', 'RT [min]', 'Annotation Level',
        'ChemSpider ID', 'mzCloud ID', 'Mass List ID', 'General Classification', 'Sub-class', 'Comments', 'Structure'
    ]
    
    output_df = output_df[desired_cols]

    #========================================================================#
    #=====                    Creating the Excel Sheet                  =====#
    #========================================================================#

    # download output_df as an excel file
    output_df.to_excel('output.xlsx', index=False)

    # Load the workbook and select the active sheet
    workbook = load_workbook("output.xlsx")
    sheet = workbook.active
    sheet.title = "Validation"

    # Define the new colors for each level
    color_map = {
        3: PatternFill(start_color="B7E1CD", end_color="B7E1CD", fill_type="solid"),  # Light green
        2: PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),  # Light yellow
        1: PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid"),  # Pinkish red
        0: PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),  # Light gray
    }

    # Find the column for 'Annotation Level'
    for cell in sheet[1]:  # Assumes header is in the first row
        if cell.value == "Annotation Level":
            annotation_column = cell.column_letter
            break

    # Apply conditional formatting rules for each level
    for level, fill in color_map.items():
        rule = CellIsRule(operator="equal", formula=[str(level)], fill=fill)
        sheet.conditional_formatting.add(f"{annotation_column}2:{annotation_column}{sheet.max_row}", rule)

    color_map = {
        "Metabolite": PatternFill(start_color="93C47D", end_color="93C47D", fill_type="solid"),  # Medium green
        "Unmatched": PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid"),   # Medium yellow
        "Synthetic": PatternFill(start_color="E6B8AF", end_color="E6B8AF", fill_type="solid"),   # Darker pinkish red
    }

    # Find the column for 'Main Category'
    for cell in sheet[1]:  # Assumes header is in the first row
        if cell.value == "Main Category":
            category_column = cell.column_letter
            break

    # Apply conditional formatting rules for each category
    for category, fill in color_map.items():
        # Using FormulaRule to check if the cell value matches the category
        rule = FormulaRule(formula=[f'ISNUMBER(SEARCH("{category}", {category_column}2))'], fill=fill)
        sheet.conditional_formatting.add(f"{category_column}2:{category_column}{sheet.max_row}", rule)

    # Find the column for 'ChemSpider ID'
    for cell in sheet[1]:  # Assumes header is in the first row
        if cell.value == "ChemSpider ID":
            chemspider_column = cell.column_letter
            break

    # Convert each non-blank cell in the 'ChemSpider ID' column to a hyperlink
    for row in range(2, sheet.max_row + 1):  # Start from row 2, assuming row 1 is the header
        cell = sheet[f"{chemspider_column}{row}"]
        if cell.value:  # If the cell is not blank
            chemspider_id = str(cell.value)
            link = f"https://www.chemspider.com/Chemical-Structure.{chemspider_id}.html"
            cell.hyperlink = link
            cell.value = int(chemspider_id)  # Keeps the text the same as the original ID
            cell.style = "Hyperlink"  # Apply hyperlink style for consistent formatting

    # Find the column for 'mzCloud ID'
    for cell in sheet[1]:  # Assumes header is in the first row
        if cell.value == "mzCloud ID":
            mzCloud_column = cell.column_letter
            break

    # Convert each non-blank cell in the 'mzCloud ID' column to a hyperlink
    for row in range(2, sheet.max_row + 1):  # Start from row 2, assuming row 1 is the header
        cell = sheet[f"{mzCloud_column}{row}"]
        if cell.value:  # If the cell is not blank
            mzCloud_id = str(cell.value)
            link = f"https://www.mzcloud.org/compound/reference/{mzCloud_id}"
            cell.hyperlink = link
            cell.value = int(mzCloud_id)  # Keeps the text the same as the original ID
            cell.style = "Hyperlink"  # Apply hyperlink style for consistent formatting

    # Find the column for 'Mass List ID'
    for cell in sheet[1]:  # Assumes header is in the first row
        if cell.value == "Mass List ID":
            masslist_column = cell.column_letter
            break

    # Convert each non-blank cell in the 'Mass List ID' column to a hyperlink
    for row in range(2, sheet.max_row + 1):  # Start from row 2, assuming row 1 is the header
        cell = sheet[f"{masslist_column}{row}"]
        if cell.value:  # If the cell is not blank
            masslist_id = str(cell.value)
            if masslist_id.startswith('FL'):
                link = f"http://metabolomics.jp/wiki/{masslist_id}"
            else:
                link = f"https://www.npatlas.org/explore/compounds/{masslist_id}"
            cell.hyperlink = link
            cell.value = masslist_id  # Keeps the text the same as the original ID
            cell.style = "Hyperlink"  # Apply hyperlink style for consistent formatting

    # Set specific column widths
    column_widths = {
        'A': 15,
        'B': 15,
        'C': 18,
        'G': 18,
        'H': 15,
        'I': 15,
        'J': 18,
        'K': 20,
        'L': 18,
        'M': 15
    }

    for column, width in column_widths.items():
        sheet.column_dimensions[column].width = width

    # Define the fills for the alternating colors
    light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Light Blue
    light_yellow_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")  # Light Yellow

    # Format the column titles
    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=1, column=col)
        cell.font = Font(bold=True)  # Make text bold

        # Apply alternating fills
        if col % 2 == 0:
            cell.fill = light_yellow_fill
        else:
            cell.fill = light_blue_fill

    # Center the title for column B
    sheet.cell(row=1, column=2).alignment = Alignment(horizontal='center')
    sheet.cell(row=1, column=11).alignment = Alignment(horizontal='center')
    sheet.cell(row=1, column=12).alignment = Alignment(horizontal='center')
    sheet.cell(row=1, column=13).alignment = Alignment(horizontal='center')

    # Center all cells in the entire worksheet, except for column B's entries
    for row in sheet.iter_rows():
        for cell in row:
            if cell.column in [2, 11, 12, 13] and cell.row > 1:  # For column B entries (skip header)
                cell.alignment = Alignment(horizontal='left')  # Left-align column B entries
            else:
                cell.alignment = Alignment(horizontal='center')  # Center align all other cells

    new_sheet = workbook.create_sheet("Structures")
    for i, cell in enumerate(sheet['N'], start=1):
        new_sheet[f'A{i}'] = cell.value  # Copy value to new sheet

    sheet.delete_cols(14)

    return workbook



    